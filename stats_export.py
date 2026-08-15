"""
stats_export.py
-----------------
Export des statistiques d'usage de ClarIA (historique de requêtes
quantifié, voir analytics.py) au format CSV et Excel (.xlsx), pour analyse
dans un tableur ou un outil de reporting d'entreprise.

Indépendant de Streamlit : testable et réutilisable tel quel.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Sequence, Union

from analytics import QueryLogEntry, aggregate_usage_stats

PathLike = Union[str, Path]

CSV_FIELDNAMES = [
    "timestamp",
    "question",
    "answer",
    "confidence_label",
    "confidence_score",
    "source_documents",
    "response_time_seconds",
    "used_web_fallback",
    "search_mode",
    "prompt_tokens_estimate",
    "completion_tokens_estimate",
    "is_answer_found",
]


def query_log_to_csv_bytes(entries: Sequence[QueryLogEntry]) -> bytes:
    """Sérialise l'historique de requêtes en CSV (en mémoire), prêt pour un
    bouton de téléchargement Streamlit. Encode en UTF-8 avec BOM pour un
    affichage correct des accents dans Excel."""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=CSV_FIELDNAMES)
    writer.writeheader()
    for entry in entries:
        writer.writerow(entry.to_dict())
    return buffer.getvalue().encode("utf-8-sig")


def export_query_log_to_csv(entries: Sequence[QueryLogEntry], output_path: PathLike) -> Path:
    """Écrit l'historique de requêtes dans un fichier CSV. Renvoie le chemin du fichier."""
    output_path = Path(output_path)
    output_path.write_bytes(query_log_to_csv_bytes(entries))
    return output_path


def export_query_log_to_excel(entries: Sequence[QueryLogEntry], output_path: PathLike) -> Path:
    """Génère un classeur Excel (.xlsx) avec deux feuilles :
      - 'Detail des requetes' : une ligne par question posée.
      - 'Synthese' : statistiques agrégées (taux de réponses sourcées,
        confiance moyenne, documents les plus consultés, etc.), prête pour
        un usage de reporting en entreprise.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path)
    workbook = Workbook()

    header_fill = PatternFill(start_color="4C6FFF", end_color="4C6FFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # -- Feuille 1 : détail des requêtes -----------------------------------
    detail_ws = workbook.active
    detail_ws.title = "Detail des requetes"
    detail_ws.append(CSV_FIELDNAMES)
    for cell in detail_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for entry in entries:
        row = entry.to_dict()
        detail_ws.append([row[field] for field in CSV_FIELDNAMES])

    for i, field in enumerate(CSV_FIELDNAMES, start=1):
        detail_ws.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(field) + 6))

    # -- Feuille 2 : synthèse ------------------------------------------------
    summary_ws = workbook.create_sheet("Synthese")
    stats = aggregate_usage_stats(entries)

    summary_rows = [
        ("Nombre total de questions", stats["total_questions"]),
        ("Taux de réponses sourcées", f"{stats['answered_rate'] * 100:.1f} %"),
        ("Taux \"information non trouvée\"", f"{stats['no_info_rate'] * 100:.1f} %"),
        ("Temps de réponse moyen (secondes)", stats["avg_response_time_seconds"]),
        ("Score de confiance moyen", stats["avg_confidence_score"]),
        ("Réponses de confiance Forte", stats["confidence_distribution"].get("Fort", 0)),
        ("Réponses de confiance Moyenne", stats["confidence_distribution"].get("Moyen", 0)),
        ("Réponses de confiance Faible", stats["confidence_distribution"].get("Faible", 0)),
        ("Réponses sans source", stats["confidence_distribution"].get("Aucune source", 0)),
        ("Recours à la recherche web", stats["web_fallback_count"]),
        ("Appels API estimés", stats["total_api_calls"]),
        ("Tokens estimés cumulés", stats["total_tokens_estimate"]),
    ]

    summary_ws.append(["Indicateur", "Valeur"])
    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in summary_rows:
        summary_ws.append(row)

    summary_ws.append([])
    summary_ws.append(["Documents les plus consultés", "Réponses citant ce document"])
    header_row_index = summary_ws.max_row
    for cell in summary_ws[header_row_index]:
        cell.fill = header_fill
        cell.font = header_font
    for doc, count in stats["most_consulted_documents"]:
        summary_ws.append([doc, count])

    summary_ws.column_dimensions["A"].width = 42
    summary_ws.column_dimensions["B"].width = 22

    workbook.save(str(output_path))
    return output_path
