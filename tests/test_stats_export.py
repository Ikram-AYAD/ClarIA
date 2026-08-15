"""
tests/test_stats_export.py
-----------------------------
Tests unitaires pour stats_export.py : export CSV (en mémoire et fichier)
et export Excel (.xlsx) de l'historique de requêtes quantifié.
"""

import csv
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import analytics  # noqa: E402
import stats_export  # noqa: E402


def _sample_entries():
    return [
        analytics.build_query_log_entry(
            question="Quel est le budget ?",
            answer="Le budget est de 4.2 millions d'euros.",
            retrieved=[],
            response_time_seconds=1.2,
            search_mode="hybrid",
        ),
        analytics.build_query_log_entry(
            question="Et les risques ?",
            answer=analytics.NO_INFO_PHRASE,
            retrieved=[],
            response_time_seconds=0.6,
            search_mode="hybrid",
        ),
    ]


# --------------------------------------------------------------------------
# CSV
# --------------------------------------------------------------------------

def test_query_log_to_csv_bytes_is_valid_csv():
    entries = _sample_entries()
    raw = stats_export.query_log_to_csv_bytes(entries)
    assert raw.startswith(b"\xef\xbb\xbf")  # BOM UTF-8

    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    assert len(rows) == 2
    assert rows[0]["question"] == "Quel est le budget ?"
    assert set(stats_export.CSV_FIELDNAMES) <= set(reader.fieldnames)


def test_export_query_log_to_csv_writes_file(tmp_path):
    entries = _sample_entries()
    output_path = tmp_path / "stats.csv"
    result_path = stats_export.export_query_log_to_csv(entries, output_path)

    assert result_path == output_path
    assert output_path.exists()
    content = output_path.read_text(encoding="utf-8-sig")
    assert "budget" in content.lower()


def test_csv_export_handles_empty_entries():
    raw = stats_export.query_log_to_csv_bytes([])
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    assert list(reader) == []
    assert set(stats_export.CSV_FIELDNAMES) <= set(reader.fieldnames)


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------

def test_export_query_log_to_excel_creates_valid_workbook(tmp_path):
    from openpyxl import load_workbook

    entries = _sample_entries()
    output_path = tmp_path / "stats.xlsx"
    result_path = stats_export.export_query_log_to_excel(entries, output_path)

    assert result_path == output_path
    assert output_path.exists()

    workbook = load_workbook(str(output_path))
    assert "Detail des requetes" in workbook.sheetnames
    assert "Synthese" in workbook.sheetnames

    detail_ws = workbook["Detail des requetes"]
    # +1 pour la ligne d'en-tête
    assert detail_ws.max_row == len(entries) + 1
    header = [cell.value for cell in detail_ws[1]]
    assert header == stats_export.CSV_FIELDNAMES


def test_export_query_log_to_excel_summary_sheet_has_stats(tmp_path):
    from openpyxl import load_workbook

    entries = _sample_entries()
    output_path = tmp_path / "stats.xlsx"
    stats_export.export_query_log_to_excel(entries, output_path)

    workbook = load_workbook(str(output_path))
    summary_ws = workbook["Synthese"]
    summary_values = [row[0].value for row in summary_ws.iter_rows()]

    assert "Nombre total de questions" in summary_values
    assert "Taux de réponses sourcées" in summary_values


def test_export_query_log_to_excel_handles_empty_entries(tmp_path):
    output_path = tmp_path / "empty.xlsx"
    result_path = stats_export.export_query_log_to_excel([], output_path)
    assert result_path.exists()
